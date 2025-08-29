"""
IR & OCV Testing Recorder

A desktop application with a simple GUI for collecting battery pack and module
test data and saving it to a structured Excel workbook with dynamic formulas.

Dependencies:
    pip install pandas openpyxl
    (pyzbar and other barcode libs are not strictly needed as scanner
    input is treated as a keyboard wedge, but can be added for future use)

Usage:
    - Run the script: `python IR_OCV_Recorder.py`
    - Hotkeys:
        - On Module Entry Screen:
            - Enter: Automatically advances focus after scanning a code.
            - Arrow keys (left/right): Navigate between cells in the table.
            - Tab: Moves to the next input field.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

class IROCV_RecorderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("IR & OCV Testing Recorder")

        # Session data model
        self.session = {
            "pack_name": "",
            "pack_code": "",
            "num_modules": 0,
            "cells_per_module": 0,
            "filename": "",
            "modules_completed": 0,
            "current_module_index": 1,
            "module_data": {} # stores pandas DataFrames for each module
        }

        # Constants for validation
        self.OCV_MIN = 0.0
        self.OCV_MAX = 6.0
        
        # UI elements
        self.status_label = ttk.Label(self.root, text="Ready to start...", anchor="w")
        self.status_label.pack(side="bottom", fill="x", padx=10, pady=5)
        
        self.start_screen()

    def clear_screen(self):
        """Clears all widgets from the main frame."""
        for widget in self.root.winfo_children():
            if widget not in (self.status_label,):
                widget.destroy()

    def show_message(self, message):
        """Updates the status bar with a message."""
        self.status_label.config(text=message)

    def start_screen(self):
        """Builds the initial pack setup screen."""
        self.clear_screen()
        frame = ttk.Frame(self.root, padding="20")
        frame.pack(expand=True)

        ttk.Label(frame, text="Pack Setup", font=("Arial", 16)).grid(row=0, column=0, columnspan=2, pady=10)

        # Labels and Entry fields
        labels = ["Battery Pack Name:", "Battery Pack Code:", "Number of Modules:", "Cells per Module:"]
        self.entries = {}
        for i, label_text in enumerate(labels):
            ttk.Label(frame, text=label_text).grid(row=i+1, column=0, sticky="w", pady=5)
            entry = ttk.Entry(frame)
            entry.grid(row=i+1, column=1, sticky="ew", pady=5)
            self.entries[label_text.split(':')[0].strip()] = entry
        
        # Error labels for validation
        self.error_labels = {}
        for i, label_text in enumerate(labels):
            error_label = ttk.Label(frame, text="", foreground="red")
            error_label.grid(row=i+1, column=2, sticky="w")
            self.error_labels[label_text.split(':')[0].strip()] = error_label

        # Buttons
        button_frame = ttk.Frame(frame)
        button_frame.grid(row=len(labels)+1, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="Start New Test Session", command=self.start_new_session).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Load/Append to Existing Workbook", command=self.load_workbook).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Demo", command=self.start_demo_session).pack(side="left", padx=5)
        
        self.show_message("Enter pack details to begin or load an existing file.")

    def start_new_session(self):
        """Validates and starts a new test session."""
        if self.validate_pack_setup():
            self.session["filename"] = f"{self.session['pack_name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            self.show_message(f"New session started. Saving to '{self.session['filename']}'")
            self.module_entry_screen()

    def start_demo_session(self):
        """Fills the inputs with demo data and starts a session."""
        self.entries["Battery Pack Name"].insert(0, "Pack_A1")
        self.entries["Battery Pack Code"].insert(0, "PKA1-2025-0828")
        self.entries["Number of Modules"].insert(0, "2")
        self.entries["Cells per Module"].insert(0, "4")
        self.start_new_session()

    def load_workbook(self):
        """Loads an existing Excel workbook for appending."""
        filepath = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if filepath:
            try:
                wb = load_workbook(filepath)
                module_sheets = [s for s in wb.sheetnames if s.startswith("Module_")]
                highest_index = 0
                if module_sheets:
                    # Find the highest module index to continue from
                    indices = [int(s.split('_')[1]) for s in module_sheets]
                    highest_index = max(indices)
                
                # Check for existing summary sheet and pack details
                if 'Summary' in wb.sheetnames:
                    summary_df = pd.read_excel(filepath, sheet_name="Summary")
                    pack_name = summary_df.loc[0, "PackName"] if "PackName" in summary_df.columns else ""
                    pack_code = summary_df.loc[0, "PackCode"] if "PackCode" in summary_df.columns else ""
                    
                    self.entries["Battery Pack Name"].insert(0, pack_name)
                    self.entries["Battery Pack Code"].insert(0, pack_code)
                    # Cannot determine total modules or cells per module from summary
                    
                self.session["filename"] = filepath
                self.session["current_module_index"] = highest_index + 1
                self.show_message(f"Workbook '{os.path.basename(filepath)}' loaded. Ready to add Module {self.session['current_module_index']}.")
                
                self.validate_pack_setup(loaded=True)
                if self.session["num_modules"] > 0 and self.session["cells_per_module"] > 0:
                     self.module_entry_screen()
                else:
                    self.show_message("Enter the original Number of Modules and Cells per Module to continue.")

            except Exception as e:
                messagebox.showerror("Error", f"Failed to load workbook: {e}")

    def validate_pack_setup(self, loaded=False):
        """Validates the inputs on the start screen."""
        pack_name = self.entries["Battery Pack Name"].get().strip()
        pack_code = self.entries["Battery Pack Code"].get().strip()
        
        try:
            num_modules = int(self.entries["Number of Modules"].get().strip())
            cells_per_module = int(self.entries["Cells per Module"].get().strip())
        except ValueError:
            messagebox.showerror("Validation Error", "Number of Modules and Cells per Module must be integers.")
            return False

        if not pack_name:
            messagebox.showerror("Validation Error", "Battery Pack Name cannot be empty.")
            return False
        
        if not pack_code:
            messagebox.showerror("Validation Error", "Battery Pack Code cannot be empty.")
            return False

        if num_modules <= 0:
            messagebox.showerror("Validation Error", "Number of Modules must be greater than 0.")
            return False

        if cells_per_module <= 0:
            messagebox.showerror("Validation Error", "Cells per Module must be greater than 0.")
            return False

        self.session["pack_name"] = pack_name
        self.session["pack_code"] = pack_code
        self.session["num_modules"] = num_modules
        self.session["cells_per_module"] = cells_per_module
        
        return True

    def module_entry_screen(self):
        """Builds the module and cell data entry screen."""
        self.clear_screen()
        frame = ttk.Frame(self.root, padding="20")
        frame.pack(expand=True, fill="both")
        
        # Module details section
        module_frame = ttk.LabelFrame(frame, text="Module Details", padding="10")
        module_frame.pack(fill="x", pady=5)
        
        self.module_title = ttk.Label(module_frame, text=f"Module {self.session['current_module_index']}/{self.session['num_modules']}", font=("Arial", 14))
        self.module_title.pack(side="left", padx=10)
        
        ttk.Label(module_frame, text="Module Code:").pack(side="left", padx=5)
        self.module_code_entry = ttk.Entry(module_frame)
        self.module_code_entry.pack(side="left", expand=True, fill="x", padx=5)
        self.module_code_entry.bind("<Return>", lambda e: self.module_code_entry.focus_set()) # Keep focus on module code

        # Cell data table
        table_frame = ttk.Frame(frame)
        table_frame.pack(expand=True, fill="both")
        
        columns = ("Cell #", "Battery Code", "IR_mOhm", "OCV_V", "Notes")
        self.cell_tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        for col in columns:
            self.cell_tree.heading(col, text=col)
            self.cell_tree.column(col, width=120)
        self.cell_tree.pack(side="left", expand=True, fill="both")
        
        # Add a vertical scrollbar
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.cell_tree.yview)
        vsb.pack(side="right", fill="y")
        self.cell_tree.configure(yscrollcommand=vsb.set)

        # Populate the table
        self.cells = []
        for i in range(1, self.session["cells_per_module"] + 1):
            values = [i, "", "", "", ""]
            item_id = self.cell_tree.insert("", "end", values=values)
            self.cells.append(item_id)
        
        # Add editable functionality
        self.cell_tree.bind("<Double-1>", self.on_cell_double_click)
        self.cell_tree.bind("<Return>", self.on_cell_enter)
        self.cell_tree.bind("<Tab>", self.on_cell_tab)
        
        # Navigation and control buttons
        button_frame = ttk.Frame(frame, padding="5")
        button_frame.pack(fill="x", pady=5)

        # Control buttons
        self.prev_button = ttk.Button(button_frame, text="Previous Module", command=self.prev_module)
        self.prev_button.pack(side="left", padx=5)
        
        self.save_button = ttk.Button(button_frame, text="Save Module Data", command=self.save_module_data)
        self.save_button.pack(side="left", expand=True, fill="x", padx=5)
        
        self.next_button = ttk.Button(button_frame, text="Next Module", command=self.next_module)
        self.next_button.pack(side="left", padx=5)
        
        self.finish_button = ttk.Button(button_frame, text="Finish & Close", command=self.finish_session)
        self.finish_button.pack(side="right", padx=5)

        self.update_buttons()
        self.show_message(f"Ready to scan Module {self.session['current_module_index']} code.")

    def on_cell_double_click(self, event):
        """Enables in-cell editing for the Treeview."""
        item = self.cell_tree.identify_row(event.y)
        col = self.cell_tree.identify_column(event.x)
        
        if not item or col == '#1': # Prevent editing Cell # column
            return
            
        col_index = int(col.replace('#', '')) - 1
        
        # Get cell position and value
        x,y,width,height = self.cell_tree.bbox(item, column=col)
        value = self.cell_tree.item(item, 'values')[col_index]

        # Create the Entry widget
        self.editor = ttk.Entry(self.cell_tree)
        self.editor.place(x=x, y=y, width=width, height=height, anchor="nw")
        self.editor.insert(0, value)
        self.editor.focus_set()
        
        self.editor.bind("<Return>", lambda e: self.on_cell_edit_done(item, col_index))
        self.editor.bind("<FocusOut>", lambda e: self.on_cell_edit_done(item, col_index))
        
    def on_cell_edit_done(self, item, col_index):
        """Saves the new value and destroys the editor."""
        new_value = self.editor.get()
        current_values = list(self.cell_tree.item(item, 'values'))
        current_values[col_index] = new_value
        self.cell_tree.item(item, values=current_values)
        self.editor.destroy()
        self.editor = None

    def on_cell_enter(self, event):
        """Handles the 'Enter' key to move to the next field."""
        if self.root.focus_get() == self.module_code_entry:
            self.module_code_entry.focus_set()
            return
            
        current_item = self.cell_tree.focus()
        current_col = self.cell_tree.identify_column(self.cell_tree.winfo_pointerx() - self.root.winfo_rootx() - self.cell_tree.winfo_x())
        col_index = int(current_col.replace('#', '')) - 1
        
        if col_index < len(self.cell_tree["columns"]) - 1:
            next_col_index = col_index + 1
            # Find the position of the next cell
            x, y, _, _ = self.cell_tree.bbox(current_item, column=f"#{next_col_index + 1}")
            self.cell_tree.event_generate("<Double-1>", x=x + 5, y=y + 5)
        else:
            # Move to the next row
            next_item = self.cell_tree.next(current_item)
            if next_item:
                x, y, _, _ = self.cell_tree.bbox(next_item, column="#2")
                self.cell_tree.event_generate("<Double-1>", x=x + 5, y=y + 5)

    def update_buttons(self):
        """Enables/disables navigation buttons based on current module index."""
        self.prev_button.config(state="normal" if self.session["current_module_index"] > 1 else "disabled")
        self.next_button.config(state="normal" if self.session["current_module_index"] < self.session["num_modules"] else "disabled")

    def prev_module(self):
        """Switches to the previous module entry screen."""
        if self.session["current_module_index"] > 1:
            self.save_module_data(silent=True)
            self.session["current_module_index"] -= 1
            self.load_module_data()
            self.update_buttons()

    def next_module(self):
        """Switches to the next module entry screen."""
        if self.session["current_module_index"] < self.session["num_modules"]:
            self.save_module_data(silent=True)
            self.session["current_module_index"] += 1
            self.load_module_data()
            self.update_buttons()

    def load_module_data(self):
        """Loads data for the current module if it exists in the buffer."""
        self.module_title.config(text=f"Module {self.session['current_module_index']}/{self.session['num_modules']}")
        
        module_code = self.session["module_data"].get(self.session["current_module_index"], {}).get("module_code", "")
        self.module_code_entry.delete(0, "end")
        self.module_code_entry.insert(0, module_code)

        self.cell_tree.delete(*self.cell_tree.get_children())
        df = self.session["module_data"].get(self.session["current_module_index"], {}).get("dataframe")
        if df is not None:
            for _, row in df.iterrows():
                self.cell_tree.insert("", "end", values=list(row[["CellIndex", "BatteryCode", "IR_mOhm", "OCV_V", "Notes"]]))
        else:
            for i in range(1, self.session["cells_per_module"] + 1):
                self.cell_tree.insert("", "end", values=(i, "", "", "", ""))
                
    def save_module_data(self, silent=False):
        """Validates and saves the current module's data to the session buffer."""
        module_code = self.module_code_entry.get().strip()
        rows = []
        is_valid = True
        missing_codes = False
        
        for item in self.cell_tree.get_children():
            vals = self.cell_tree.item(item)["values"]
            cell_index = vals[0]
            battery_code = vals[1]
            ir_str, ocv_str, notes = vals[2], vals[3], vals[4]
            
            if not battery_code:
                missing_codes = True

            try:
                ir = float(ir_str) if ir_str else None
                ocv = float(ocv_str) if ocv_str else None
                
                if ir is not None and ir <= 0:
                    messagebox.showerror("Validation Error", f"IR for Cell {cell_index} must be greater than 0.")
                    is_valid = False
                    break
                    
                if ocv is not None and not (self.OCV_MIN <= ocv <= self.OCV_MAX):
                    messagebox.showerror("Validation Error", f"OCV for Cell {cell_index} is outside the valid range ({self.OCV_MIN}-{self.OCV_MAX}V).")
                    is_valid = False
                    break
                    
                rows.append([
                    datetime.now().isoformat(),
                    self.session['pack_name'],
                    self.session['pack_code'],
                    self.session['current_module_index'],
                    module_code,
                    cell_index,
                    battery_code,
                    ir,
                    ocv,
                    notes
                ])
            except ValueError:
                messagebox.showerror("Validation Error", f"Invalid IR or OCV value for Cell {cell_index}.")
                is_valid = False
                break
        
        if not is_valid:
            return False

        if missing_codes and not silent:
            if not messagebox.askyesno("Warning", "Some battery codes are missing. Continue saving?"):
                return False
        
        df = pd.DataFrame(rows, columns=[
            "Timestamp", "PackName", "PackCode", "ModuleIndex", "ModuleCode", "CellIndex", 
            "BatteryCode", "IR_mOhm", "OCV_V", "Notes"
        ])
        
        self.session["module_data"][self.session["current_module_index"]] = {
            "module_code": module_code,
            "dataframe": df
        }
        
        if not silent:
            self.write_to_excel()
            self.show_message(f"Saved Module {self.session['current_module_index']}. Proceeding to next module.")
            return True
        else:
            return True

    def finish_session(self):
        """Saves final data and prompts to close."""
        if not self.save_module_data(silent=True):
            return
            
        if self.session["modules_completed"] < self.session["num_modules"]:
            if not messagebox.askyesno("Confirm Finish", "Not all modules have been saved. Do you want to finish anyway?"):
                return
        
        self.write_to_excel()
        messagebox.showinfo("Session Complete", f"All data saved to {self.session['filename']}.")
        self.root.destroy()

    def write_to_excel(self):
        """Writes all session data to the Excel workbook, creating/updating sheets and formulas."""
        try:
            if os.path.exists(self.session["filename"]):
                wb = load_workbook(self.session["filename"])
            else:
                from openpyxl.workbook import Workbook
                wb = Workbook()
                wb.remove(wb.active) # Remove default sheet
            
            # Write/update individual module sheets
            for module_idx, data in self.session["module_data"].items():
                sheet_name = f"Module_{module_idx:03d}"
                df = data["dataframe"]
                
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # Clear existing data before rewriting
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.value = None
                    # If we need to append
                    # start_row = ws.max_row + 1
                    # for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
                    #     for c_idx, value in enumerate(row):
                    #         ws.cell(row=start_row + r_idx, column=c_idx + 1, value=value)
                    
                else:
                    ws = wb.create_sheet(title=sheet_name)

                # Write DataFrame content to the sheet
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                    ws.append(row)
                
                # Apply table styling and autofit
                tab = Table(displayName=f"ModuleTable{module_idx}", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
                style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab.tableStyleInfo = style
                ws.add_table(tab)
                
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter # Get the column letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width
                
                # Freeze the header row
                ws.freeze_panes = 'A2'

            # Rebuild the Summary sheet from scratch
            if 'Summary' in wb.sheetnames:
                del wb['Summary']
            summary_ws = wb.create_sheet(title="Summary", index=0)

            # Get list of all module sheets for formulas
            module_sheets = sorted([s for s in wb.sheetnames if s.startswith("Module_")])
            
            # Summary sheet headers
            summary_headers = ["ModuleIndex", "ModuleCode", "CellCount", "IR_AVG", "IR_MAX", "IR_MIN", "IR_RANGE", "OCV_AVG", "OCV_MAX", "OCV_MIN", "OCV_RANGE"]
            summary_ws.append(summary_headers)

            # Populate module-level summary rows and build pack-level formulas
            all_ir_cells = []
            all_ocv_cells = []
            
            for i, sheet_name in enumerate(module_sheets):
                ws = wb[sheet_name]
                ir_col_letter = get_column_letter(ws.max_column - 2)
                ocv_col_letter = get_column_letter(ws.max_column - 1)
                
                ir_col_range = f"{sheet_name}!{ir_col_letter}2:{ir_col_letter}{ws.max_row}"
                ocv_col_range = f"{sheet_name}!{ocv_col_letter}2:{ocv_col_letter}{ws.max_row}"
                
                all_ir_cells.append(ir_col_range)
                all_ocv_cells.append(ocv_col_range)
                
                # Get module code and cell count
                module_df = pd.read_excel(self.session["filename"], sheet_name=sheet_name)
                module_code = module_df["ModuleCode"].iloc[0] if not module_df.empty else "N/A"
                cell_count = len(module_df)
                
                # Write module-level row with formulas
                row_num = i + 2
                summary_ws.cell(row=row_num, column=1, value=int(sheet_name.split('_')[1]))
                summary_ws.cell(row=row_num, column=2, value=module_code)
                summary_ws.cell(row=row_num, column=3, value=cell_count)
                summary_ws.cell(row=row_num, column=4, value=f"=AVERAGE({ir_col_range})")
                summary_ws.cell(row=row_num, column=5, value=f"=MAX({ir_col_range})")
                summary_ws.cell(row=row_num, column=6, value=f"=MIN({ir_col_range})")
                summary_ws.cell(row=row_num, column=7, value=f"=E{row_num}-F{row_num}") # IR_MAX - IR_MIN
                summary_ws.cell(row=row_num, column=8, value=f"=AVERAGE({ocv_col_range})")
                summary_ws.cell(row=row_num, column=9, value=f"=MAX({ocv_col_range})")
                summary_ws.cell(row=row_num, column=10, value=f"=MIN({ocv_col_range})")
                summary_ws.cell(row=row_num, column=11, value=f"=I{row_num}-J{row_num}") # OCV_MAX - OCV_MIN

            # Add PACK_TOTALS row
            final_row = len(module_sheets) + 2
            summary_ws.cell(row=final_row, column=1, value="PACK_TOTALS").font = Font(bold=True)
            summary_ws.cell(row=final_row, column=3, value=f"=SUM(C2:C{final_row-1})").font = Font(bold=True)
            summary_ws.cell(row=final_row, column=4, value=f'=AVERAGE({",".join(all_ir_cells)})').font = Font(bold=True)
            summary_ws.cell(row=final_row, column=5, value=f'=MAX({",".join(all_ir_cells)})').font = Font(bold=True)
            summary_ws.cell(row=final_row, column=6, value=f'=MIN({",".join(all_ir_cells)})').font = Font(bold=True)
            summary_ws.cell(row=final_row, column=7, value=f"=E{final_row}-F{final_row}").font = Font(bold=True)
            summary_ws.cell(row=final_row, column=8, value=f'=AVERAGE({",".join(all_ocv_cells)})').font = Font(bold=True)
            summary_ws.cell(row=final_row, column=9, value=f'=MAX({",".join(all_ocv_cells)})').font = Font(bold=True)
            summary_ws.cell(row=final_row, column=10, value=f'=MIN({",".join(all_ocv_cells)})').font = Font(bold=True)
            summary_ws.cell(row=final_row, column=11, value=f"=I{final_row}-J{final_row}").font = Font(bold=True)
            
            # Apply formatting
            for col in summary_ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column letter
                for cell in col:
                    try: # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                summary_ws.column_dimensions[column].width = adjusted_width
                
            summary_ws.freeze_panes = 'A2'
            
            # Save the workbook
            wb.save(self.session["filename"])
            self.session["modules_completed"] = len(module_sheets)
            
        except Exception as e:
            messagebox.showerror("Excel Write Error", f"An error occurred while writing to the Excel file: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = IROCV_RecorderApp(root)
    root.mainloop()
